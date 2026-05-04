#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import subprocess
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Best-known solutions para ta01-ta80 do repositório thomasWeise/jsspInstancesAndResults
BEST_KNOWN_SOLUTIONS = {
    "ta01": 1231, "ta02": 1244, "ta03": 1218, "ta04": 1175, "ta05": 1224,
    "ta06": 1238, "ta07": 1227, "ta08": 1217, "ta09": 1274, "ta10": 1241,
    "ta11": 1357, "ta12": 1367, "ta13": 1342, "ta14": 1345, "ta15": 1339,
    "ta16": 1360, "ta17": 1462, "ta18": 1396, "ta19": 1332, "ta20": 1348,
    "ta21": 1642, "ta22": 1600, "ta23": 1557, "ta24": 1644, "ta25": 1595,
    "ta26": 1643, "ta27": 1680, "ta28": 1603, "ta29": 1625, "ta30": 1584,
    "ta31": 1764, "ta32": 1784, "ta33": 1791, "ta34": 1829, "ta35": 2007,
    "ta36": 1819, "ta37": 1771, "ta38": 1673, "ta39": 1795, "ta40": 1669,
    "ta41": 2005, "ta42": 1937, "ta43": 1846, "ta44": 1979, "ta45": 2000,
    "ta46": 2004, "ta47": 1889, "ta48": 1937, "ta49": 1961, "ta50": 1923,
    "ta51": 2760, "ta52": 2756, "ta53": 2717, "ta54": 2839, "ta55": 2679,
    "ta56": 2781, "ta57": 2943, "ta58": 2885, "ta59": 2655, "ta60": 2723,
    "ta61": 2868, "ta62": 2869, "ta63": 2755, "ta64": 2702, "ta65": 2725,
    "ta66": 2845, "ta67": 2825, "ta68": 2784, "ta69": 3071, "ta70": 2995,
    "ta71": 5464, "ta72": 5181, "ta73": 5568, "ta74": 5339, "ta75": 5392,
    "ta76": 5342, "ta77": 5436, "ta78": 5394, "ta79": 5358, "ta80": 5183,
}

def extract_makespan(output):
    """
    Extrai o makespan da saída do programa.
    Procura por: "Makespan (Comprimento do caminho maximo): XXX"
    """
    match = re.search(r'Makespan \(Comprimento do caminho maximo\): (\d+)', output)
    if match:
        return int(match.group(1))
    return None

def run_instance(executable, instance_path):
    """
    Executa o programa para uma instância e retorna o makespan.
    """
    try:
        result = subprocess.run(
            [executable, instance_path],
            capture_output=True,
            text=True,
            timeout=60
        )
        
        if result.returncode == 0:
            makespan = extract_makespan(result.stdout)
            return makespan
        else:
            print(f"Erro ao executar {instance_path}: {result.stderr}")
            return None
    except subprocess.TimeoutExpired:
        print(f"Timeout ao executar {instance_path}")
        return None
    except Exception as e:
        print(f"Exceção ao executar {instance_path}: {e}")
        return None

def main():
    # Configurações
    base_dir = os.path.dirname(os.path.abspath(__file__))
    exe_name = "jobshop.exe" if os.name == "nt" else "jobshop"
    executable = os.path.join(base_dir, exe_name)
    instances_dir = os.path.join(base_dir, "Instancias", "JSP")
    
    # Verifica se o executável existe
    if not os.path.exists(executable):
        print(f"Executável não encontrado: {executable}")
        print("Compilando o programa...")
        result = subprocess.run(["make"], cwd=base_dir, capture_output=True, text=True)
        if result.returncode != 0:
            print(f"Erro ao compilar: {result.stderr}")
            return
    
    # Verifica se o diretório de instâncias existe
    if not os.path.exists(instances_dir):
        print(f"Diretório de instâncias não encontrado: {instances_dir}")
        return
    
    # Coleta resultados
    results = []
    for i in range(1, 81):
        instance_name = f"ta{i:02d}Jsp.psi"
        instance_key = f"ta{i:02d}"
        instance_path = os.path.join(instances_dir, instance_name)
        
        if not os.path.exists(instance_path):
            print(f"Instância não encontrada: {instance_name}")
            results.append((instance_name, None, None, None))
            continue
        
        print(f"Processando {instance_name}...", end=" ")
        makespan = run_instance(executable, instance_path)
        
        if makespan is not None:
            bks = BEST_KNOWN_SOLUTIONS.get(instance_key)
            gap = makespan - bks if bks else None
            gap_percentage = (gap / bks * 100) if bks and gap else None
            
            print(f"Makespan: {makespan}, BKS: {bks}, Gap: {gap}")
            results.append((instance_name, makespan, bks, gap, gap_percentage))
        else:
            print(f"Erro ao processar")
            results.append((instance_name, None, None, None, None))
    
    # Cria arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Adiciona cabeçalhos
    headers = ["Instância", "Makespan Obtido", "Best Known Solution", "Gap", "Gap (%)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Adiciona dados
    for row_idx, result in enumerate(results, start=2):
        if len(result) == 5:
            instance_name, makespan, bks, gap, gap_pct = result
        else:
            instance_name, makespan, bks, gap, gap_pct = result + (None,) * (5 - len(result))
        
        ws.cell(row=row_idx, column=1).value = instance_name
        ws.cell(row=row_idx, column=2).value = makespan if makespan is not None else "Erro"
        ws.cell(row=row_idx, column=3).value = bks if bks is not None else "N/A"
        ws.cell(row=row_idx, column=4).value = gap if gap is not None else "N/A"
        ws.cell(row=row_idx, column=5).value = f"{gap_pct:.2f}%" if gap_pct is not None else "N/A"
        
        # Centraliza os valores
        for col in range(1, 6):
            ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
    
    # Ajusta largura das colunas
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    
    # Salva arquivo
    output_file = os.path.join(base_dir, "resultados_ta01_ta80.xlsx")
    wb.save(output_file)
    print(f"\nResultados salvos em: {output_file}")

if __name__ == "__main__":
    main()
